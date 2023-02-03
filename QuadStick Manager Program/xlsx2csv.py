#-------------------------------------------------------------------------------
# Name:        xlsx2csv
# Purpose:     Convert QuadStick configuration spreadsheets from XLST format
#              to the CSV format used by the QuadStick
#
# Author:      fdavison
#
# Created:     13/09/2015
# Copyright:   (c) fdavison 2015
# Licence:     <your licence>
#-------------------------------------------------------------------------------
import sys, os
import urllib.request, urllib.error, urllib.parse, urllib.request, urllib.parse, urllib.error
import tempfile
import shutil
from openpyxl import load_workbook
import pyrfc6266
import time

def get_google_doc_from_id(id):
    try: # to get published spreadsheet
        url = "https://docs.google.com/spreadsheets/d/" + id + "/export?format=xlsx"
        print("URL= ", url)
        f = urllib.request.urlopen(url)
        x = f.read()
        if x[0:2] != b'PK': raise Exception("Spreadsheet not shared publicly")
        #print repr(dir(f))
        print(repr(f.geturl()))
        print(repr(f.info()))
        print(repr(f.getcode()))
        print(repr(dir(f.headers)))
    except Exception as e:
        print(repr(e))
        try: # to get publicly shared spreadsheet
            url = "https://docs.google.com/spreadsheets/d/" + id + "/pub?output=xlsx"
            print("URL= ", url)
            f = urllib.request.urlopen(url)
            x = f.read()
            if x[0:2] != b'PK': raise Exception("Spreadsheet not published")
        except Exception as e:
            print(repr(e))
            return None, None
            # the next level would be to log in using google api services
    name = None
    try:
        cd = dict(f.getheaders())
        cd = cd.get('Content-Disposition')
        print ("CONTENT DISPOSITION ", repr(cd))
        #hd = f.info().dict
        #cd = hd.get("content-disposition")
        # parse a string like: 'Content-Disposition: attachment; filename="AndroidMouse.xlsx"; filename*=UTF-8\'\'Android%20Mouse.xlsx\r\n'
        ph = pyrfc6266.parse_filename(cd)
        print ("parsed filename ", repr(ph))
        name = ph #.filename_sanitized('xlsx')
    except Exception as e:
        print ('content disposition exception', repr(e))
        pass
    #print(repr(x), name)
    return x, name

def get_id_from_url(url):
    # take the url from a google spreadsheet edit or published window
    if url.find("google.com") > 0 and url.find("/spreadsheets/") > 0:
        return url.split("/d/")[1].split("/")[0]
    # try next method...
    if url.find("/") < 0: #assume it is an ID number
        return url
    return None
    
def get_name_from_csv(f,d):
    try:
        with open(d + "\\" + f, "r", 0) as csv_file:
            x = csv_file.readline()
            x = x.split(",")
            if x[0] == "QuadStick Configuration":
                version = x[1].split(" ")[1] # get version number 1.4, 1.5, etc
                if float(version) > 1.3:
                    return x[3].strip()
    except Exception as e:
        print(repr(e))
    return ""

def write_temporary_file(name, contents, delete_old=True):
    tmp_folder_path = tempfile.gettempdir() + '\\quad_stick_temporary_files'
    tmp_file_path = tmp_folder_path + "\\" + name
    if delete_old:
        shutil.rmtree(tmp_folder_path, True)  # clear out any old files from before
        time.sleep(2)
        tmp_folder = os.mkdir(tmp_folder_path)
    # write zip file to temp folder and unzip it
    if type(contents) == str:
        contents = contents.encode()
    with open(tmp_file_path, "wb", 0) as xlsxFile:
        xlsxFile.write(contents)
        xlsxFile.flush()
        os.fsync(xlsxFile.fileno())
    return tmp_file_path

def get_config_profile_info(id):
    #open the file passed in with the command
    x, name = get_google_doc_from_id(id)
    if x is None:
        print ("Share or Publish Spreadsheet")
        return None
    if name is None:
        name = "temp.xlsx"
    filename = write_temporary_file(name, x)
    print(filename)
    #read the xlsx file into a Workbook object
    wb = load_workbook(filename)
    #get csv file name from cell A2 of first page
    csv_file_name = "config.csv" # default name
    ws = wb.active
    if ws['A2'].value:
        csv_file_name = str(ws['A2'].value)
    # return data in dict object form to be stored in QMP persistant data file
    return {"name":str(name[:-5]), "id": str(id), "csv_name":str(csv_file_name)}, wb

def get_config_profile_info_from_url(url):
    id = get_id_from_url(url)
    if id:
        answer = get_config_profile_info(id)
        if answer: return answer[0]
    return None

MAXCOL = 10

def write_csv_file_for(id, drive=None, QMP=None):
    #open a CSV file for output
    #convert the Workbook into CSV and write it to the file
    #close file and exit.
    info, wb = get_config_profile_info(id)
    csv_file_name = info.get("csv_name")
    if drive:
        csv_file_name = drive + csv_file_name
    id = info.get("id")
    name = info.get("name")
    print("write_csv_file_for: ", id, name, csv_file_name)
    
    lines = []
    
    lines.append('QuadStick Configuration,Version 1.5,%s,%s' % (id,name))
    for ws in wb:
        name = ws.title
        if name == 'Inputs' or name == 'Outputs' or name == "Reference Card" or name == "Voice":
            continue
        sheet_type = ws['A1'].value
        if not (sheet_type.find('Profile') >= 0 or sheet_type == 'Preferences' or sheet_type == 'Infrared'):
            print("Cell A1 does not contain valid value: ", repr(sheet_type))
            if QMP:
                QMP.text_ctrl_messages.AppendText('**ERROR** Sheet: ' + str(name) + ' Cell A1 value is invalid: ' + repr(sheet_type) + '\n')
            continue
        i = 0
        for row in ws.rows:
            i = i + 1
            if (row[0].value is not None) or (i < 4): # stop at first empty row, after header
                s = ""
                for cell in row[:MAXCOL]: # only first MAXCOL columns are used
                    if cell.value is None:
                        s = s + ","
                    else:
                        value = cell.value
                        if cell.data_type == 'n': # convert numeric types to integers
                            value = int(value)
                        s = s + str(value) + ","
                lines.append(s)
        lines.append("")
    if drive:
        with open(csv_file_name, 'w') as csv_file:
            csv_file.write("\n".join(lines)) # blank line separates sheets in csv file
            csv_file.flush()
            os.fsync(csv_file.fileno())
    else: # try serial port
        from microterm import microterm
        mt = microterm()
        mt.write_qs_file(csv_file_name, lines)
    return True


def main():
    #filename = sys.argv[1]
    url = sys.argv[1]
    id = get_id_from_url(url)
    write_csv_file_for(id)

if __name__ == '__main__':
    main()
